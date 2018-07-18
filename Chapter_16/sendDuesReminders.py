#! python3
# sendDuesRemingerd.py - Sends emails based on payment status in spreadsheet

import openpyxl,smtplib

#email
import os, sys
os.chdir(r'D:\Drive\Code\ATBSWP\Chapter_16')
sys.path.insert(0, r'D:\Drive\Code\Random')
import file
myEmail = file.email

# Open the spreadsheet and get the latest dues status. 
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']

lastCol = sheet.max_column
latestMonth = sheet.cell(row=1,column=lastCol).value
# Check each member's payment status.
unpaidMembers = {}
for r in range(2,sheet.max_row + 1):
    payment = sheet.cell(row=r,column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r,column=1).value
        email = sheet.cell(row=r,column=2).value
        unpaidMembers[name] = email

# Log in to email account.
smtpObj = smtplib.SMTP('smtp.gmail.com',587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login(myEmail,input())

# Send out reminder emails.
for name, email in unpaidMembers.items():
    body = f"Subject: {latestMonth} dues unpaid.\nDear {name},\nRecords show that you have not paid dues for {latestMonth}. Please make this paymen as soon as possible. Thank you!"
    print(f'Sending email to {email}')
    sendmailStatus = smtpObj.sendmail(myEmail,email,body)

    if sendmailStatus != {}:
        print(f'There was a problemn sending email to {email}: {sendmailStatus}')
smtpObj.quit()

