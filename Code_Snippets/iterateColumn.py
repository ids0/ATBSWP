import openpyxl, os
from openpyxl.utils.cell import get_column_letter, column_index_from_string
os.chdir('D:/Drive/Code/ATBSWP/Chapter_12')
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
columnB = list(list(sheet.columns)[1])
for cell in columnB:
    print(cell.value)
for rowNum in range(2, sheet.max_row):  # skip the first row
    sheet.cell(row=rowNum, column=1).value
input()



