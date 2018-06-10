import openpyxl, os, pprint
from openpyxl.utils.cell import get_column_letter, column_index_from_string
os.chdir('D:/Drive/Code/ATBSWP/Chapter_12')
wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb.get_sheet_by_name('Sheet1')
sheet = wb['Sheet1']
# pprint.pprint(list(sheet['A1':'C3']))
for row in sheet.iter_rows('B{}:B{}'.format(ws.min_row,ws.max_row)):
    for cell in row:
        print(cell.value)

input()
