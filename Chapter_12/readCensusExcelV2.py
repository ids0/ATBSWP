#! python3
# readCensusExcelV2.py -

import openpyxl, os,pprint
os.chdir('D:/Drive/Code/ATBSWP/Chapter_12')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
ws = wb[wb.sheetnames[0]]
counties = list(list(ws.columns)[2])
countiesDic  = {}
# county = ws.cell(row=[2:Final], column=3).value
# countyValue = ws.cell(row=[2:Final], colum=4).value
for i in range(2,ws.max_row):
    countiesDic.setdefault(ws.cell(row=i,column=3).value,0)
    countiesDic[ws.cell(row=i,column=3).value] += ws.cell(row=i,column=4).value

# Acces cell by index, locate 



pprint.pprint(countiesDic)
input()