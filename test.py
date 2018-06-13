import os, openpyxl
from openpyxl.styles import Font
os.chdir('D:/Drive/Code/ATBSWP/Chapter_12')
wb = openpyxl.Workbook()
sheet = wb['Sheet']
italic24Font = Font(size=24, italic=True)

fontObj1 = Font(name='Times New Roman',bold=True)
sheet['A1'].font = fontObj1
sheet['A1'] = 'Bold Times New Roman'

fontObj2 = Font(size=24, italic=True)
sheet['B3'].font = fontObj2
sheet['B3'] = '24 pt Italic'

sheet['A1'] = 'Tall  row'
sheet['A2'] = 300
sheet['A3'] = '=SUM(A2:A2)'
sheet['B2'] = 'Wide Column'

# sheet.row_dimensions[1].height = 70
# sheet.column_dimensions['B'].width =50

sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merge together'
sheet['D3'] = 'Twelve cells merge together?!!!'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells.'
sheet['D5'] = 'Hidden cell?'
sheet.unmerge_cells('C5:D5')


wb.save('styled.xlsx')

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
sheet.freeze_panes = 'A2'
wb.save('freezeExample.xlsx')
