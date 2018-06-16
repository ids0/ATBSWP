#! python3
# blankRowInserter.py - Insert blank rows
import os, openpyxl,sys
os.chdir('D:/Drive/Code/ATBSWP/Chapter_12')
from openpyxl.utils.cell import get_column_letter, column_index_from_string

# if len(sys.argv) > 1:
while True:
    wbName = 'test.xlsx'
    N = 3
    M = 2
    startRow = N
    blankRows = M
    # startRow = sys.argv[1]
    # blankRows = sys.argv[2]
    # wbName = sys.argv[3]
    wb = openpyxl.load_workbook(wbName)
    ws = wb.active
    print('Working')
    maxRow = ws.max_row
    maxCol = ws.max_column
    for rowNum in range(maxRow, startRow-1, -1):
         for colNum in range(1,maxCol):
            cellFrom = ws.cell(row=rowNum, column=colNum)
            cellTo = ws.cell(row=(rowNum+blankRows), column=colNum)
            print('%s to %s'%(cellFrom.coordinate,cellTo.coordinate))
            print('Cell From value %s - Cell To value %s'%(cellFrom.value,cellTo.value))
            print('----')
            cellTo.value = cellFrom.value
            # cellTo.value = 'test'
    for rowNum in range(startRow,startRow+blankRows):
        for colNum in range(1,maxCol):
            cellFrom = ws.cell(row=rowNum, column=colNum).value = ''
    wb.save('testV2.xlsx')
    break
print('Done')
input()