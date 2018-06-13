#! python3
# updateProduce.py - Corrects costos in produce sales spreadsheet.

import openpyxl,os
os.chdir('D:/Drive/Code/ATBSWP/Chapter_12')
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# The produce types and ther updated prices
PRICE_UPDATES = {'Garlic':  3.07,
                 'Celery':  1.19,
                 'Lemon':   1.27}

# Loop through the ros and update the prices
print('Updating price...')
for rowNum in range(2, sheet.max_row):  # skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
print('Done, saving changes.')

wb.save('UpdatedProduceSales.xlsx')
