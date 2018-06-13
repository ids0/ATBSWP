#! python3
# multiplicationTable.py -
import os, openpyxl, sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
os.chdir('D:/Drive/Code/ATBSWP/Chapter_12')

# if len(sys.argv) > 1:
while True:
    # N = sys.argv[1]
    N = 9
    wb = openpyxl.Workbook()
    ws = wb.active
    for rowNum in range(1,N+2):
        for colNum in range(1,N+2):
            if rowNum == 1:
                if colNum != 1:
                    ws.cell(row = rowNum, column = colNum).value = colNum-1
                    ws.cell(row = rowNum, column = colNum).font = Font(bold=True)
            if colNum == 1:
                if rowNum != 1:
                    ws.cell(row = rowNum, column = colNum).value = rowNum-1
                    ws.cell(row = rowNum, column = colNum).font = Font(bold=True)
            if rowNum != 1 and colNum != 1:
                ws.cell(row = rowNum, column = colNum).value = '=A%s*%s1' %(rowNum,get_column_letter(colNum))
    break
wb.save('multiplicationTable.xlsx')
print('Done')
input()