#! python3
# cellInverter.py - 

import openpyxl,os
os.chdir('D:/Drive/Code/ATBSWP/Chapter_12')
print('Opening workbook')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
ws = wb.active
newWb = openpyxl.Workbook()
newWs = newWb.active
print('Inverting worksheet')
for row in range(1,ws.max_row):
    for col in range(1,ws.max_column):
        newWs.cell(row=col, column=row).value = ws.cell(row=row, column=col).value
        


newWb.save('testV3.xlsx')
print('Done')
input()